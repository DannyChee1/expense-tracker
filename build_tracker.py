"""Build/refresh expense_tracker.xlsx from every CSV and PDF in imports/."""

import json
import re
import sys
from pathlib import Path
from datetime import datetime

import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

BASE_DIR = Path(__file__).parent
IMPORTS_DIR = BASE_DIR / "imports"
CATEGORIES_FILE = BASE_DIR / "categories.json"
OUTPUT_FILE = BASE_DIR / "expense_tracker.xlsx"

NO_CITY = "(no city)"
PAYMENTS_CAT = "Payments & Credits"
AMEX_MERCHANT_WIDTH = 23

SCHEMA = ["SourceFile", "Card", "Date", "PostDate", "Description",
          "City", "Merchant", "Amount", "PreCategory"]

MERCHANT_MAP_FILE = BASE_DIR / "merchant_map.json"

MONTHS = {m: i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1)}
PROVINCES = "ON|QC|BC|AB|MB|SK|NS|NB|NL|PE|NT|YT|NU"

DATE = r"[A-Z][a-z]{2} \d{2}"
CIBC_CHARGE_RE = re.compile(
    rf"^(?P<trans>{DATE}) (?P<post>{DATE}) (?P<mid>.+) "
    rf"(?P<prov>{PROVINCES}) [A-Za-z][A-Za-z ]+? "
    rf"(?P<amt>-?[\d,]+\.\d{{2}})(?P<cr> CR)?$")
CIBC_PAY_RE = re.compile(
    rf"^(?P<trans>{DATE}) (?P<post>{DATE}) (?P<desc>.+) "
    rf"(?P<amt>-?[\d,]+\.\d{{2}})(?P<cr> CR)?$")
CIBC_PERIOD_RE = re.compile(
    r"from ([A-Z][a-z]+) (\d+) to ([A-Z][a-z]+) (\d+), (\d{4})")

_STORE_NUM = re.compile(r"#\s*\d+")
_LONG_NUM = re.compile(r"\b\d{3,}\b")
_REF_AFTER_SLASH = re.compile(r"/.*$")
_POS_PREFIX = re.compile(r"^(TST|SQ|SQU|EB|SP)\s*[-*]\s*", re.I)


def normalize_city(city):
    return city.title() if city else NO_CITY


def clean_merchant(raw, city):
    m = _POS_PREFIX.sub("", raw)
    m = _STORE_NUM.sub(" ", m)
    m = _REF_AFTER_SLASH.sub("", m)
    m = _LONG_NUM.sub(" ", m)
    if city:
        m = re.sub(re.escape(city) + r"\s*$", "", m, flags=re.I)
    m = re.sub(r"\s+", " ", m).strip(" -*")
    return smart_title(m or raw)


def smart_title(text):
    titled = text.title()
    return re.sub(r"(?<=')[A-Z](?= |$)", lambda m: m.group().lower(), titled)


def split_amex_merchant_city(description):
    if len(description) > AMEX_MERCHANT_WIDTH and description[AMEX_MERCHANT_WIDTH] == " ":
        return description[:AMEX_MERCHANT_WIDTH].strip(), (description[AMEX_MERCHANT_WIDTH + 1:].strip() or None)
    return description.strip(), None


def parse_amex_csv(path):
    raw = pd.read_csv(path)
    cols = {c.lower().strip(): c for c in raw.columns}

    def find(*cands, avoid=None):
        for cand in cands:
            for lower, orig in cols.items():
                if cand in lower and (avoid is None or avoid not in lower):
                    return orig
        return None

    date_col = find("date", avoid="process")
    post_col = find("process", "posted") or date_col
    desc_col = find("description", "merchant", "details", "payee")
    amount_col = find("amount")
    debit_col, credit_col = find("debit"), find("credit")

    dates = pd.to_datetime(raw[date_col], errors="coerce")
    posts = pd.to_datetime(raw[post_col], errors="coerce")
    if amount_col:
        amounts = pd.to_numeric(raw[amount_col], errors="coerce")
    else:
        debit = pd.to_numeric(raw[debit_col], errors="coerce").fillna(0) if debit_col else 0
        credit = pd.to_numeric(raw[credit_col], errors="coerce").fillna(0) if credit_col else 0
        amounts = debit - credit

    rows = []
    for date, post, desc, amt in zip(dates, posts, raw[desc_col].astype(str), amounts):
        if pd.isna(date) or pd.isna(amt):
            continue
        merchant, city = split_amex_merchant_city(desc.strip())
        rows.append([path.name, "Amex", date, post if pd.notna(post) else date,
                     merchant, normalize_city(city), clean_merchant(merchant, city),
                     round(float(amt), 2), None])
    return pd.DataFrame(rows, columns=SCHEMA)


def pdf_lines(path):
    lines = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            lines.extend((page.extract_text() or "").splitlines())
    return lines


def statement_years(lines):
    for line in lines:
        m = CIBC_PERIOD_RE.search(line)
        if m:
            start_m, _, end_m, _, year = m.groups()
            return MONTHS[start_m[:3]], MONTHS[end_m[:3]], int(year)
    return 1, 12, datetime.now().year


def cibc_date(token, start_month, end_month, end_year):
    mon, day = token.split()
    month = MONTHS[mon]
    if start_month <= end_month:
        year = end_year
    else:
        year = end_year if month <= end_month else end_year - 1
    return datetime(year, month, int(day))


def parse_cibc_pdf(path):
    lines = pdf_lines(path)
    start_m, end_m, end_y = statement_years(lines)

    def to_amount(amt, cr):
        value = float(amt.replace(",", ""))
        return -value if cr else value

    rows, mode = [], None
    for line in lines:
        if "Your payments" in line:
            mode = "pay"
            continue
        if "Your new charges and credits" in line:
            mode = "charge"
            continue
        if line.startswith("Total payments") or "Information about" in line:
            mode = None
            continue
        if mode is None:
            continue

        charge = CIBC_CHARGE_RE.match(line)
        if mode == "charge" and charge:
            mid, city = charge["mid"].rsplit(" ", 1)
            amount = to_amount(charge["amt"], charge["cr"])
            rows.append([path.name, "CIBC",
                         cibc_date(charge["trans"], start_m, end_m, end_y),
                         cibc_date(charge["post"], start_m, end_m, end_y),
                         mid.strip(), normalize_city(city),
                         clean_merchant(mid.strip(), city), round(amount, 2), None])
            continue

        pay = CIBC_PAY_RE.match(line)
        if pay:
            amount = to_amount(pay["amt"], pay["cr"])
            desc = pay["desc"].strip()
            is_payment = mode == "pay"
            rows.append([path.name, "CIBC",
                         cibc_date(pay["trans"], start_m, end_m, end_y),
                         cibc_date(pay["post"], start_m, end_m, end_y),
                         desc, NO_CITY, clean_merchant(desc, None),
                         round(-abs(amount) if is_payment else amount, 2),
                         PAYMENTS_CAT if is_payment else None])
    return pd.DataFrame(rows, columns=SCHEMA)


def load_all():
    files = sorted(p for p in IMPORTS_DIR.iterdir()
                   if p.suffix.lower() in (".csv", ".pdf") and not p.name.startswith("~"))
    if not files:
        raise FileNotFoundError(f"No CSV or PDF files found in {IMPORTS_DIR}.")

    frames, skipped = [], []
    for path in files:
        try:
            if path.suffix.lower() == ".pdf":
                text = "".join(pdf_lines(path))
                if "CIBC" not in text:
                    skipped.append(f"{path.name} (unrecognized PDF - only CIBC supported)")
                    continue
                frame = parse_cibc_pdf(path)
            else:
                frame = parse_amex_csv(path)
            if len(frame):
                frames.append(frame)
            else:
                skipped.append(f"{path.name} (no transactions found)")
        except Exception as exc:
            skipped.append(f"{path.name} ({exc})")

    for note in skipped:
        print(f"  ! skipped {note}")
    if not frames:
        raise ValueError("No transactions could be parsed from imports/.")
    return pd.concat(frames, ignore_index=True)


def dedupe(df):
    key = ["Card", "Date", "PostDate", "Description", "Amount"]
    occ = df.groupby(["SourceFile"] + key).cumcount()
    before = len(df)
    df = df.assign(_occ=occ).drop_duplicates(subset=key + ["_occ"])
    removed = before - len(df)
    return df.drop(columns=["_occ", "SourceFile"]).reset_index(drop=True), removed


def load_merchant_map():
    if MERCHANT_MAP_FILE.exists():
        try:
            return {k.strip().lower(): v for k, v in json.loads(MERCHANT_MAP_FILE.read_text()).items()}
        except (ValueError, AttributeError):
            return {}
    return {}


def save_merchant_map(mmap):
    MERCHANT_MAP_FILE.write_text(json.dumps(dict(sorted(mmap.items())), indent=2))


def keyword_category(description, rules):
    low = description.lower()
    for category, keywords in rules.items():
        if any(kw.lower() in low for kw in keywords):
            return category
    return "Uncategorized"


def enrich_dates(df):
    df = df.copy()
    df["Month"] = df["Date"].dt.strftime("%Y-%m")
    df["DayOfWeek"] = df["Date"].dt.strftime("%a")
    return df.sort_values("Date").reset_index(drop=True)


def apply_categories(df, rules, mmap):
    categories = []
    for pre, merchant, desc in zip(df["PreCategory"], df["Merchant"], df["Description"]):
        if isinstance(pre, str) and pre:
            categories.append(pre)
        elif merchant.strip().lower() in mmap:
            categories.append(mmap[merchant.strip().lower()])
        else:
            categories.append(keyword_category(desc, rules))
    out = df.copy()
    out["Category"] = categories
    return out


def prompt_new_merchants(df, rules, mmap):
    if not sys.stdin.isatty():
        return mmap
    unc = df[df["Category"] == "Uncategorized"].copy()
    if unc.empty:
        return mmap
    unc["key"] = unc["Merchant"].str.strip().str.lower()
    spend = unc.groupby("key")["Amount"].sum()
    todo = sorted((k for k in unc["key"].unique() if k not in mmap),
                  key=lambda k: -spend[k])
    if not todo:
        return mmap

    menu = list(rules.keys())
    for value in dict.fromkeys(mmap.values()):
        if value not in menu and value != "Uncategorized":
            menu.append(value)

    print(f"\n{len(todo)} new merchant(s) to categorize.")
    print("Type a number, a NEW category name, Enter to skip, or 'q' to stop.\n")
    for i, name in enumerate(menu, 1):
        print(f"  {i}) {name}")
    print()

    for n, key in enumerate(todo, 1):
        rows = unc[unc["key"] == key]
        label = f"[{n}/{len(todo)}] {rows['Merchant'].iloc[0]}  ({len(rows)} txns, ${rows['Amount'].sum():,.2f})  > "
        while True:
            choice = input(label).strip()
            if choice.lower() == "q":
                save_merchant_map(mmap)
                return mmap
            if choice == "":
                break
            if choice.isdigit():
                idx = int(choice)
                if 1 <= idx <= len(menu):
                    mmap[key] = menu[idx - 1]
                    save_merchant_map(mmap)
                    break
                print("  number out of range")
                continue
            mmap[key] = choice
            if choice not in menu:
                menu.append(choice)
                print(f"  added new category: {choice}")
            save_merchant_map(mmap)
            break
    return mmap


MONEY = '$#,##0.00;($#,##0.00);-'
INK = "1F4E78"
HEADER_FILL = PatternFill("solid", fgColor=INK)
TILE_FILL = PatternFill("solid", fgColor="EEF3F9")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", bold=True, size=20, color=INK)
SUB_FONT = Font(name="Arial", italic=True, size=9, color="808080")
TILE_LABEL = Font(name="Arial", size=9, color="5A5A5A")
TILE_VALUE = Font(name="Arial", size=18, bold=True, color=INK)
BODY = Font(name="Arial")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DAY_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

TX_COLS = ["Date", "Card", "Merchant", "Description", "Category", "City",
           "Amount", "Month", "DayOfWeek"]


def write_transactions(ws, df):
    for i, head in enumerate(TX_COLS, start=1):
        c = ws.cell(1, i, head)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    for r, row in enumerate(df.itertuples(index=False), start=2):
        ws.cell(r, 1, row.Date.strftime("%Y-%m-%d")).font = BODY
        ws.cell(r, 2, row.Card).font = BODY
        ws.cell(r, 3, row.Merchant).font = BODY
        ws.cell(r, 4, row.Description).font = BODY
        ws.cell(r, 5, row.Category).font = BODY
        ws.cell(r, 6, row.City).font = BODY
        amt = ws.cell(r, 7, round(float(row.Amount), 2))
        amt.number_format = MONEY
        amt.font = BODY
        ws.cell(r, 8, row.Month).font = BODY
        ws.cell(r, 9, row.DayOfWeek).font = BODY
    for i, w in enumerate([12, 8, 26, 34, 18, 16, 13, 9, 6], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(df) + 1}"


def write_value_table(ws, top, title, series, money=True):
    ws.cell(top, 1, title).font = Font(name="Arial", bold=True)
    header = top + 1
    for col, label in ((1, "Item"), (2, "Amount")):
        c = ws.cell(header, col, label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for i, (label, value) in enumerate(series.items(), start=1):
        r = header + i
        ws.cell(r, 1, str(label)).font = BODY
        c = ws.cell(r, 2, round(float(value), 2))
        c.font = BODY
        if money:
            c.number_format = MONEY
        ws.cell(r, 1).border = BORDER
        ws.cell(r, 2).border = BORDER
    return header, header + len(series)


def chart_from(kind, title, ws, header, last, *, percent=False, values=False,
               bar_dir="bar", height=8, width=15):
    chart = kind()
    chart.title = title
    chart.style = 10
    data = Reference(ws, min_col=2, min_row=header, max_row=last)
    cats = Reference(ws, min_col=1, min_row=header + 1, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height, chart.width = height, width
    if kind is BarChart:
        chart.type = bar_dir
        chart.gapWidth = 60
        chart.legend = None
    if kind is LineChart:
        chart.legend = None
    if percent or values:
        labels = DataLabelList()
        labels.showPercent = percent
        labels.showVal = values
        labels.numFmt = MONEY if values else None
        chart.dataLabels = labels
    return chart


def tile(ws, col, label, value, money=False):
    span = f"{get_column_letter(col)}5:{get_column_letter(col + 2)}5"
    vspan = f"{get_column_letter(col)}6:{get_column_letter(col + 2)}6"
    ws.merge_cells(span)
    ws.merge_cells(vspan)
    lab = ws.cell(5, col, label)
    lab.font = TILE_LABEL
    lab.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    val = ws.cell(6, col, value)
    val.font = TILE_VALUE
    val.alignment = Alignment(horizontal="left", vertical="center")
    if money:
        val.number_format = MONEY
    for c in range(col, col + 3):
        ws.cell(5, c).fill = TILE_FILL
        ws.cell(6, c).fill = TILE_FILL
        ws.cell(5, c).border = BORDER
        ws.cell(6, c).border = BORDER


def build_workbook(df, dupes_removed):
    wb = Workbook()
    tx = wb.active
    tx.title = "Transactions"
    write_transactions(tx, df)

    spend = df[df["Category"] != PAYMENTS_CAT]
    purchases = spend[spend["Amount"] > 0]
    total_spend = spend["Amount"].sum()
    avg_purchase = purchases["Amount"].mean() if len(purchases) else 0
    by_cat = spend.groupby("Category")["Amount"].sum().sort_values(ascending=False)
    by_city = spend.groupby("City")["Amount"].sum().sort_values(ascending=False)
    by_card = spend.groupby("Card")["Amount"].sum().sort_values(ascending=False)
    by_dow = spend.groupby("DayOfWeek")["Amount"].sum().reindex(DAY_ORDER).fillna(0)
    by_merch = spend.groupby("Merchant")["Amount"].sum().sort_values(ascending=False).head(12)
    by_month = spend.groupby("Month")["Amount"].sum().sort_index()
    cumulative = spend.groupby("Date")["Amount"].sum().sort_index().cumsum()
    cumulative.index = [d.strftime("%m-%d") for d in cumulative.index]
    top_city = by_city.index[0] if len(by_city) else "-"
    period = f"{df['Date'].min():%b %d} - {df['Date'].max():%b %d, %Y}"

    data = wb.create_sheet("Data")
    data.column_dimensions["A"].width = 26
    data.column_dimensions["B"].width = 14
    row = 1
    tables = {}
    for name, series in (("cat", by_cat), ("city", by_city), ("card", by_card),
                         ("month", by_month), ("dow", by_dow), ("merch", by_merch),
                         ("cum", cumulative)):
        title = {"cat": "Spend by Category", "city": "Spend by City",
                 "card": "Spend by Card", "month": "Spend by Month",
                 "dow": "Spend by Day of Week", "merch": "Top Merchants",
                 "cum": "Cumulative Spend"}[name]
        tables[name] = write_value_table(data, row, title, series)
        row = tables[name][1] + 2

    dash = wb.create_sheet("Dashboard", 0)
    dash.sheet_view.showGridLines = False
    dash.column_dimensions["A"].width = 2
    for c in range(2, 20):
        dash.column_dimensions[get_column_letter(c)].width = 10

    dash.merge_cells("B2:J2")
    dash["B2"] = "Expense Dashboard"
    dash["B2"].font = TITLE_FONT
    dash.merge_cells("B3:J3")
    dash["B3"] = f"Updated {datetime.now():%Y-%m-%d %H:%M}    |    Period {period}"
    dash["B3"].font = SUB_FONT
    dash.row_dimensions[5].height = 30

    tile(dash, 2, "Total Spend (excl. payments)", round(float(total_spend), 2), money=True)
    tile(dash, 6, "# Purchases", int(len(purchases)))
    tile(dash, 10, "Avg Purchase", round(float(avg_purchase), 2), money=True)
    tile(dash, 14, "Top City", top_city)

    dash.add_chart(chart_from(BarChart, "Spend by Category", data, *tables["cat"],
                              values=True, bar_dir="bar", height=11, width=28), "B9")
    dash.add_chart(chart_from(BarChart, "Spend by Month", data, *tables["month"],
                              bar_dir="col"), "B33")
    dash.add_chart(chart_from(BarChart, "Spend by City", data, *tables["city"]), "K33")
    dash.add_chart(chart_from(BarChart, "Top Merchants", data, *tables["merch"]), "B50")
    dash.add_chart(chart_from(BarChart, "Spend by Day of Week", data, *tables["dow"],
                              bar_dir="col"), "K50")
    dash.add_chart(chart_from(PieChart, "Spend by Card", data, *tables["card"],
                              percent=True), "B67")
    dash.add_chart(chart_from(LineChart, "Cumulative Spend Over Time", data,
                              *tables["cum"], height=9, width=26), "B84")

    save(wb, len(df), dupes_removed)


def save(wb, n, dupes_removed):
    target = OUTPUT_FILE
    try:
        wb.save(target)
    except PermissionError:
        target = OUTPUT_FILE.with_name(f"expense_tracker_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
        wb.save(target)
        print(f"  ! {OUTPUT_FILE.name} was open in Excel - wrote {target.name} instead.")
    print(f"Done: {n} transactions, {dupes_removed} duplicates removed -> {target.name}")


def main():
    try:
        df = load_all()
    except (FileNotFoundError, ValueError) as exc:
        print(f"Nothing to do: {exc}")
        return 1
    df, dupes_removed = dedupe(df)
    rules = json.loads(CATEGORIES_FILE.read_text())
    mmap = load_merchant_map()
    df = enrich_dates(df)
    df = apply_categories(df, rules, mmap)
    mmap = prompt_new_merchants(df, rules, mmap)
    df = apply_categories(df, rules, mmap)
    build_workbook(df, dupes_removed)
    return 0


if __name__ == "__main__":
    sys.exit(main())
