"""Build/refresh expense_tracker.xlsx from every CSV and PDF in imports/."""

import json
import re
import sys
from pathlib import Path
from datetime import datetime

import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.workbook.defined_name import DefinedName

# This file lives in the app/ subfolder; imports and the output xlsx live one
# level up (the folder the user double-clicks in), configs live next to the code.
APP_DIR = Path(__file__).parent
ROOT_DIR = APP_DIR.parent
IMPORTS_DIR = ROOT_DIR / "imports"
OUTPUT_FILE = ROOT_DIR / "expense_tracker.xlsx"
CATEGORIES_FILE = APP_DIR / "categories.json"

NO_CITY = "(no city)"
PAYMENTS_CAT = "Payments & Credits"
AMEX_MERCHANT_WIDTH = 23

SCHEMA = ["SourceFile", "Card", "Date", "PostDate", "Description",
          "City", "Merchant", "Amount", "PreCategory"]

MERCHANT_MAP_FILE = APP_DIR / "merchant_map.json"

# ---------------------------------------------------------------------------
# Display formats - edit these to change how dates and numbers are shown.
# Date codes are Python strftime (%Y year, %b month name, %d day, %a weekday).
# Money codes are Excel number formats.
# ---------------------------------------------------------------------------
FMT_DATE_CELL = "yyyy-mm-dd"          # Date column in the Transactions tab
FMT_MONTH = "%Y-%m"                    # Month value + "Spend by Month" labels
FMT_DOW = "%a"                         # Day-of-week labels (Mon, Tue, ...)
FMT_DAILY_LABEL = "%b\n%d"             # daily spend-over-time x-axis (month over day)
FMT_PERIOD = "%b %d"                   # period start/end in the subtitle
FMT_UPDATED = "%Y-%m-%d %H:%M"         # "Updated ..." timestamp
MONEY = '$#,##0.00;($#,##0.00);-'      # money format for cells
AXIS_MONEY = '"$"#,##0'                # money format for chart axes/labels

MONTHS = {m: i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], start=1)}
PROVINCES = "ON|QC|BC|AB|MB|SK|NS|NB|NL|PE|NT|YT|NU"

DATE = r"[A-Z][a-z]{2} \d{2}"
CIBC_CHARGE_RE = re.compile(
    rf"^(?P<trans>{DATE}) (?P<post>{DATE}) (?P<mid>.+) "
    rf"(?P<prov>{PROVINCES}) [A-Za-z][A-Za-z ]+? "
    rf"(?P<amt>-?[\d,]+\.\d{{2}})(?P<cr> CR)?$")
CIBC_PAY_RE = re.compile(
    rf"^(?P<trans>{DATE}) (?P<post>{DATE}) (?P<desc>.+) "
    rf"(?P<amt>-?[\d,]+\.\d{{2}})(?P<cr> CR)?$")
CIBC_PERIOD_RE = re.compile(
    r"from ([A-Z][a-z]+) (\d+) to ([A-Z][a-z]+) (\d+), (\d{4})")

_STORE_NUM = re.compile(r"#\s*\d+")
_LONG_NUM = re.compile(r"\b\d{3,}\b")
_REF_AFTER_SLASH = re.compile(r"/.*$")
_POS_PREFIX = re.compile(r"^(TST|SQ|SQU|EB|SP)\s*[-*]\s*", re.I)


def normalize_city(city):
    return city.title() if city else NO_CITY


def clean_merchant(raw, city):
    m = _POS_PREFIX.sub("", raw)
    m = _STORE_NUM.sub(" ", m)
    m = _REF_AFTER_SLASH.sub("", m)
    m = _LONG_NUM.sub(" ", m)
    if city:
        m = re.sub(re.escape(city) + r"\s*$", "", m, flags=re.I)
    m = re.sub(r"\s+", " ", m).strip(" -*")
    return smart_title(m or raw)


def smart_title(text):
    titled = text.title()
    return re.sub(r"(?<=')[A-Z](?= |$)", lambda m: m.group().lower(), titled)


def split_amex_merchant_city(description):
    if len(description) > AMEX_MERCHANT_WIDTH and description[AMEX_MERCHANT_WIDTH] == " ":
        return description[:AMEX_MERCHANT_WIDTH].strip(), (description[AMEX_MERCHANT_WIDTH + 1:].strip() or None)
    return description.strip(), None


def parse_amex_csv(path):
    raw = pd.read_csv(path)
    cols = {c.lower().strip(): c for c in raw.columns}

    def find(*cands, avoid=None):
        for cand in cands:
            for lower, orig in cols.items():
                if cand in lower and (avoid is None or avoid not in lower):
                    return orig
        return None

    date_col = find("date", avoid="process")
    post_col = find("process", "posted") or date_col
    desc_col = find("description", "merchant", "details", "payee")
    amount_col = find("amount")
    debit_col, credit_col = find("debit"), find("credit")

    missing = [name for name, col in (("Date", date_col), ("Description", desc_col))
               if col is None]
    if amount_col is None and not (debit_col or credit_col):
        missing.append("Amount")
    if missing:
        raise ValueError(f"CSV is missing a {', '.join(missing)} column - not a supported export")

    dates = pd.to_datetime(raw[date_col], errors="coerce")
    posts = pd.to_datetime(raw[post_col], errors="coerce")
    if amount_col:
        amounts = pd.to_numeric(raw[amount_col], errors="coerce")
    else:
        debit = pd.to_numeric(raw[debit_col], errors="coerce").fillna(0) if debit_col else 0
        credit = pd.to_numeric(raw[credit_col], errors="coerce").fillna(0) if credit_col else 0
        amounts = debit - credit

    rows = []
    for date, post, desc, amt in zip(dates, posts, raw[desc_col].astype(str), amounts):
        if pd.isna(date) or pd.isna(amt):
            continue
        merchant, city = split_amex_merchant_city(desc.strip())
        rows.append([path.name, "Amex", date, post if pd.notna(post) else date,
                     merchant, normalize_city(city), clean_merchant(merchant, city),
                     round(float(amt), 2), None])
    return pd.DataFrame(rows, columns=SCHEMA)


def pdf_lines(path):
    lines = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            lines.extend((page.extract_text() or "").splitlines())
    return lines


def statement_years(lines):
    for line in lines:
        m = CIBC_PERIOD_RE.search(line)
        if m:
            start_m, _, end_m, _, year = m.groups()
            return MONTHS[start_m[:3]], MONTHS[end_m[:3]], int(year)
    return 1, 12, datetime.now().year


def cibc_date(token, start_month, end_month, end_year):
    mon, day = token.split()
    month = MONTHS[mon]
    if start_month <= end_month:
        year = end_year
    else:
        year = end_year if month <= end_month else end_year - 1
    return datetime(year, month, int(day))


def parse_cibc_pdf(path):
    lines = pdf_lines(path)
    start_m, end_m, end_y = statement_years(lines)

    def to_amount(amt, cr):
        value = float(amt.replace(",", ""))
        return -value if cr else value

    rows, mode = [], None
    for line in lines:
        if "Your payments" in line:
            mode = "pay"
            continue
        if "Your new charges and credits" in line:
            mode = "charge"
            continue
        if line.startswith("Total payments") or "Information about" in line:
            mode = None
            continue
        if mode is None:
            continue

        charge = CIBC_CHARGE_RE.match(line)
        if mode == "charge" and charge:
            mid, city = charge["mid"].rsplit(" ", 1)
            amount = to_amount(charge["amt"], charge["cr"])
            rows.append([path.name, "CIBC",
                         cibc_date(charge["trans"], start_m, end_m, end_y),
                         cibc_date(charge["post"], start_m, end_m, end_y),
                         mid.strip(), normalize_city(city),
                         clean_merchant(mid.strip(), city), round(amount, 2), None])
            continue

        pay = CIBC_PAY_RE.match(line)
        if pay:
            amount = to_amount(pay["amt"], pay["cr"])
            desc = pay["desc"].strip()
            is_payment = mode == "pay"
            rows.append([path.name, "CIBC",
                         cibc_date(pay["trans"], start_m, end_m, end_y),
                         cibc_date(pay["post"], start_m, end_m, end_y),
                         desc, NO_CITY, clean_merchant(desc, None),
                         round(-abs(amount) if is_payment else amount, 2),
                         PAYMENTS_CAT if is_payment else None])
    return pd.DataFrame(rows, columns=SCHEMA)


def load_all():
    files = sorted(p for p in IMPORTS_DIR.iterdir()
                   if p.suffix.lower() in (".csv", ".pdf") and not p.name.startswith("~"))
    if not files:
        raise FileNotFoundError(f"No CSV or PDF files found in {IMPORTS_DIR}.")

    frames, skipped = [], []
    for path in files:
        try:
            if path.suffix.lower() == ".pdf":
                text = "".join(pdf_lines(path))
                if "CIBC" not in text:
                    skipped.append(f"{path.name} (unrecognized PDF - only CIBC supported)")
                    continue
                frame = parse_cibc_pdf(path)
            else:
                frame = parse_amex_csv(path)
            if len(frame):
                frames.append(frame)
            else:
                skipped.append(f"{path.name} (no transactions found)")
        except Exception as exc:
            skipped.append(f"{path.name} ({exc})")

    for note in skipped:
        print(f"  ! skipped {note}")
    if not frames:
        raise ValueError("No transactions could be parsed from imports/.")
    return pd.concat(frames, ignore_index=True)


def dedupe(df):
    key = ["Card", "Date", "PostDate", "Description", "Amount"]
    occ = df.groupby(["SourceFile"] + key).cumcount()
    before = len(df)
    df = df.assign(_occ=occ).drop_duplicates(subset=key + ["_occ"])
    removed = before - len(df)
    return df.drop(columns=["_occ", "SourceFile"]).reset_index(drop=True), removed


def load_merchant_map():
    if MERCHANT_MAP_FILE.exists():
        try:
            return {k.strip().lower(): v for k, v in json.loads(MERCHANT_MAP_FILE.read_text()).items()}
        except (ValueError, AttributeError):
            return {}
    return {}


def save_merchant_map(mmap):
    MERCHANT_MAP_FILE.write_text(json.dumps(dict(sorted(mmap.items())), indent=2))


def keyword_category(description, rules):
    low = description.lower()
    for category, keywords in rules.items():
        if any(kw.lower() in low for kw in keywords):
            return category
    return "Uncategorized"


def enrich_dates(df):
    df = df.copy()
    df["Month"] = df["Date"].dt.strftime(FMT_MONTH)
    df["DayOfWeek"] = df["Date"].dt.strftime(FMT_DOW)
    return df.sort_values("Date").reset_index(drop=True)


def apply_categories(df, rules, mmap):
    categories = []
    for pre, merchant, desc in zip(df["PreCategory"], df["Merchant"], df["Description"]):
        if isinstance(pre, str) and pre:
            categories.append(pre)
        elif merchant.strip().lower() in mmap:
            categories.append(mmap[merchant.strip().lower()])
        else:
            categories.append(keyword_category(desc, rules))
    out = df.copy()
    out["Category"] = categories
    return out


def prompt_new_merchants(df, rules, mmap):
    if not sys.stdin.isatty():
        return mmap
    unc = df[df["Category"] == "Uncategorized"].copy()
    if unc.empty:
        return mmap
    unc["key"] = unc["Merchant"].str.strip().str.lower()
    spend = unc.groupby("key")["Amount"].sum()
    todo = sorted((k for k in unc["key"].unique() if k not in mmap),
                  key=lambda k: -spend[k])
    if not todo:
        return mmap

    menu = list(rules.keys())
    for value in dict.fromkeys(mmap.values()):
        if value not in menu and value != "Uncategorized":
            menu.append(value)

    print(f"\n{len(todo)} new merchant(s) to categorize.")
    print("Type a number, a NEW category name, Enter to skip, or 'q' to stop.\n")
    for i, name in enumerate(menu, 1):
        print(f"  {i}) {name}")
    print()

    for n, key in enumerate(todo, 1):
        rows = unc[unc["key"] == key]
        label = f"[{n}/{len(todo)}] {rows['Merchant'].iloc[0]}  ({len(rows)} txns, ${rows['Amount'].sum():,.2f})  > "
        while True:
            choice = input(label).strip()
            if choice.lower() == "q":
                save_merchant_map(mmap)
                return mmap
            if choice == "":
                break
            if choice.isdigit():
                idx = int(choice)
                if 1 <= idx <= len(menu):
                    mmap[key] = menu[idx - 1]
                    save_merchant_map(mmap)
                    break
                print("  number out of range")
                continue
            mmap[key] = choice
            if choice not in menu:
                menu.append(choice)
                print(f"  added new category: {choice}")
            save_merchant_map(mmap)
            break
    return mmap


INK = "1F4E78"
HEADER_FILL = PatternFill("solid", fgColor=INK)
TILE_FILL = PatternFill("solid", fgColor="EEF3F9")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", bold=True, size=20, color=INK)
SUB_FONT = Font(name="Arial", italic=True, size=9, color="808080")
TILE_LABEL = Font(name="Arial", size=9, color="5A5A5A")
TILE_VALUE = Font(name="Arial", size=18, bold=True, color=INK)
BODY = Font(name="Arial")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DAY_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

TX_COLS = ["Date", "Card", "Merchant", "Description", "Category", "City",
           "Amount", "Month", "DayOfWeek"]


def write_transactions(ws, df):
    for i, head in enumerate(TX_COLS, start=1):
        c = ws.cell(1, i, head)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    for r, row in enumerate(df.itertuples(index=False), start=2):
        d = ws.cell(r, 1, row.Date.to_pydatetime())
        d.number_format = FMT_DATE_CELL
        d.font = BODY
        ws.cell(r, 2, row.Card).font = BODY
        ws.cell(r, 3, row.Merchant).font = BODY
        ws.cell(r, 4, row.Description).font = BODY
        ws.cell(r, 5, row.Category).font = BODY
        ws.cell(r, 6, row.City).font = BODY
        amt = ws.cell(r, 7, round(float(row.Amount), 2))
        amt.number_format = MONEY
        amt.font = BODY
        ws.cell(r, 8, row.Month).font = BODY
        ws.cell(r, 9, row.DayOfWeek).font = BODY
    for i, w in enumerate([12, 8, 26, 34, 18, 16, 13, 9, 6], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    last = len(df) + 1
    table = Table(displayName="TransactionsTable", ref=f"A1:I{last}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    return last


def _table_header(data, top, title):
    data.cell(top, 1, title).font = Font(name="Arial", bold=True)
    header = top + 1
    for col, label in ((1, "Item"), (2, "Amount")):
        c = data.cell(header, col, label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    return header


def write_dim_table(data, top, title, rows, key_range, base, amt, gross=False):
    """One breakdown table whose values are filter-aware SUMPRODUCT formulas.
    rows is a list of (display_label, match_value); match_value goes in hidden col C."""
    header = _table_header(data, top, title)
    for i, (disp, match) in enumerate(rows, start=1):
        r = header + i
        data.cell(r, 1, disp).font = BODY
        data.cell(r, 3, match)
        expr = f'({key_range}=$C${r})*{base}'
        if gross:
            expr += f'*({amt}>0)'
        cell = data.cell(r, 2)
        cell.value = f'=SUMPRODUCT({expr}*{amt})'
        cell.number_format = MONEY
        cell.font = BODY
        data.cell(r, 1).border = BORDER
        data.cell(r, 2).border = BORDER
    return header, header + len(rows)


def write_size_table(data, top, title, bins, base, amt):
    header = _table_header(data, top, title)
    for i, (lo, hi, label) in enumerate(bins, start=1):
        r = header + i
        data.cell(r, 1, label).font = BODY
        cond = f'{base}*({amt}>0)*({amt}>={lo})'
        if hi is not None:
            cond += f'*({amt}<{hi})'
        cell = data.cell(r, 2)
        cell.value = f'=SUMPRODUCT({cond})'
        cell.number_format = '0'
        cell.font = BODY
        data.cell(r, 1).border = BORDER
        data.cell(r, 2).border = BORDER
    return header, header + len(bins)


def _refs(ws, header, last):
    return (Reference(ws, min_col=2, min_row=header, max_row=last),
            Reference(ws, min_col=1, min_row=header + 1, max_row=last))


def _value_labels(show_val=False, show_pct=False):
    dl = DataLabelList()
    dl.showVal = show_val
    dl.showPercent = show_pct
    dl.showCatName = False
    dl.showSerName = False
    dl.showLegendKey = False
    dl.showBubbleSize = False
    if show_val:
        dl.numFmt = AXIS_MONEY
    return dl


def column_chart(title, ws, header, last, *, width=15, height=9):
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.style = 10
    chart.gapWidth = 40
    data, cats = _refs(ws, header, last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.numFmt = AXIS_MONEY
    chart.y_axis.majorGridlines = None
    chart.dataLabels = _value_labels(show_val=True)
    chart.width, chart.height = width, height
    return chart


def pie_chart(title, ws, header, last, *, width=15, height=9):
    chart = PieChart()
    chart.title = title
    chart.style = 10
    data, cats = _refs(ws, header, last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = _value_labels(show_pct=True)
    chart.legend.position = "r"
    chart.width, chart.height = width, height
    return chart


def line_chart(title, ws, header, last, *, width=32, height=9):
    chart = LineChart()
    chart.title = title
    chart.style = 12
    data, cats = _refs(ws, header, last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.y_axis.numFmt = AXIS_MONEY
    series = chart.series[0]
    series.smooth = False
    series.graphicalProperties.line.solidFill = INK
    series.graphicalProperties.line.width = 22000
    chart.width, chart.height = width, height
    return chart


def tile(ws, col, label, value, row=9, money=False, number=False):
    lrow, vrow = row, row + 1
    ws.merge_cells(start_row=lrow, start_column=col, end_row=lrow, end_column=col + 2)
    ws.merge_cells(start_row=vrow, start_column=col, end_row=vrow, end_column=col + 2)
    lab = ws.cell(lrow, col, label)
    lab.font = TILE_LABEL
    lab.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    val = ws.cell(vrow, col, value)
    val.font = TILE_VALUE
    val.alignment = Alignment(horizontal="left", vertical="center")
    if money:
        val.number_format = MONEY
    elif number:
        val.number_format = '0'
    for c in range(col, col + 3):
        for rr in (lrow, vrow):
            ws.cell(rr, c).fill = TILE_FILL
            ws.cell(rr, c).border = BORDER
    ws.row_dimensions[lrow].height = 16


def filter_label(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(name="Arial", bold=True, size=9, color="5A5A5A")
    ws[cell].alignment = Alignment(horizontal="right", vertical="center")


def control_cell(ws, cell, value, is_date=False):
    c = ws[cell]
    c.value = value
    c.font = Font(name="Arial", size=11, bold=True, color=INK)
    c.fill = TILE_FILL
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
    if is_date:
        c.number_format = FMT_DATE_CELL


def build_workbook(df, dupes_removed, city="(All)", category="(All)",
                   card="(All)", out=None):
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    tx = wb.active
    tx.title = "Transactions"
    n = write_transactions(tx, df)

    spend = df[df["Category"] != PAYMENTS_CAT]
    cat_labels = list(spend.groupby("Category")["Amount"].sum().sort_values(ascending=False).index)
    city_labels = list(spend.groupby("City")["Amount"].sum().sort_values(ascending=False).index)
    card_labels = list(spend.groupby("Card")["Amount"].sum().sort_values(ascending=False).index)
    month_labels = sorted(spend["Month"].unique())
    merch_full = list(spend.groupby("Merchant")["Amount"].sum().sort_values(ascending=False).head(8).index)
    merch_disp = [m if len(m) <= 12 else m[:11] + "…" for m in merch_full]

    dmin = df["Date"].min().to_pydatetime()
    dmax = df["Date"].max().to_pydatetime()
    period = f"{dmin.strftime(FMT_PERIOD)} - {dmax.strftime(FMT_PERIOD)}, {dmax.year}"

    daily_rows = [(d.strftime(FMT_DAILY_LABEL), d.to_pydatetime())
                  for d in pd.date_range(dmin, dmax, freq="D")]

    f_city, f_cat, f_card = "Dashboard!$C$5", "Dashboard!$G$5", "Dashboard!$K$5"
    f_from, f_to = "Dashboard!$C$7", "Dashboard!$G$7"

    def rng(col):
        return f"Transactions!${col}$2:${col}${n}"
    r_date, r_card, r_merch, r_cat = rng("A"), rng("B"), rng("C"), rng("E")
    r_city, r_amt, r_month, r_dow = rng("F"), rng("G"), rng("H"), rng("I")

    base = (f'(({r_city}={f_city})+({f_city}="(All)"))'
            f'*(({r_cat}={f_cat})+({f_cat}="(All)"))'
            f'*(({r_card}={f_card})+({f_card}="(All)"))'
            f'*({r_date}>={f_from})*({r_date}<={f_to})'
            f'*({r_cat}<>"{PAYMENTS_CAT}")')

    data = wb.create_sheet("Data")
    data.column_dimensions["A"].width = 22
    data.column_dimensions["B"].width = 14

    T = {}
    row = 1
    T["cat"] = write_dim_table(data, row, "Spend by Category", [(c, c) for c in cat_labels], r_cat, base, r_amt); row = T["cat"][1] + 2
    T["city"] = write_dim_table(data, row, "Spend by City", [(c, c) for c in city_labels], r_city, base, r_amt); row = T["city"][1] + 2
    T["card"] = write_dim_table(data, row, "Spend by Card", [(c, c) for c in card_labels], r_card, base, r_amt); row = T["card"][1] + 2
    T["month"] = write_dim_table(data, row, "Spend by Month", [(m, m) for m in month_labels], r_month, base, r_amt); row = T["month"][1] + 2
    T["dow"] = write_dim_table(data, row, "Spend by Day of Week", [(d, d) for d in DAY_ORDER], r_dow, base, r_amt, gross=True); row = T["dow"][1] + 2
    T["merch"] = write_dim_table(data, row, "Top Merchants", list(zip(merch_disp, merch_full)), r_merch, base, r_amt); row = T["merch"][1] + 2
    size_bins = [(0, 10, "$0-10"), (10, 25, "$10-25"), (25, 50, "$25-50"), (50, 100, "$50-100"), (100, None, "$100+")]
    T["size"] = write_size_table(data, row, "Transaction Size", size_bins, base, r_amt); row = T["size"][1] + 2

    a_header = _table_header(data, row, "Spend Over Time (daily)")
    for i, (disp, d) in enumerate(daily_rows, start=1):
        r = a_header + i
        data.cell(r, 1, disp).font = BODY
        v = data.cell(r, 2)
        v.value = f'=SUMPRODUCT(({r_date}=DATE({d.year},{d.month},{d.day}))*{base}*({r_amt}>0)*{r_amt})'
        v.number_format = MONEY
    T["trend"] = (a_header, a_header + len(daily_rows))

    for col, values in ((20, ["(All)"] + city_labels), (21, ["(All)"] + cat_labels),
                        (22, ["(All)"] + card_labels)):
        for i, v in enumerate(values, start=1):
            data.cell(i, col, v)

    for col in ("C", "T", "U", "V"):
        data.column_dimensions[col].hidden = True

    dash = wb.create_sheet("Dashboard", 0)
    dash.sheet_view.showGridLines = False
    dash.column_dimensions["A"].width = 2
    for c in range(2, 29):
        dash.column_dimensions[get_column_letter(c)].width = 10
    dash.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    dash.page_setup.orientation = "landscape"
    dash.page_setup.fitToWidth = 1
    dash.page_setup.fitToHeight = 0

    dash.merge_cells("B2:R2")
    dash["B2"] = "Expense Dashboard"
    dash["B2"].font = TITLE_FONT
    dash.merge_cells("B3:R3")
    dash["B3"] = f"Updated {datetime.now().strftime(FMT_UPDATED)}    |    Period {period}"
    dash["B3"].font = SUB_FONT

    filters = [
        ("B5", "City", "C5:D5", city, False),
        ("F5", "Category", "G5:H5", category, False),
        ("J5", "Card", "K5:L5", card, False),
        ("B7", "From", "C7:D7", dmin, True),
        ("F7", "To", "G7:H7", dmax, True),
    ]
    for label_cell, text, ctrl_range, value, is_date in filters:
        filter_label(dash, label_cell, text)
        dash.merge_cells(ctrl_range)
        control_cell(dash, ctrl_range.split(":")[0], value, is_date=is_date)
        for row_cells in dash[ctrl_range]:
            for cc in row_cells:
                cc.fill = TILE_FILL
                cc.border = BORDER

    for name, ref in (("CityList", f"Data!$T$1:$T${len(city_labels) + 1}"),
                      ("CatList", f"Data!$U$1:$U${len(cat_labels) + 1}"),
                      ("CardList", f"Data!$V$1:$V${len(card_labels) + 1}")):
        wb.defined_names[name] = DefinedName(name, attr_text=ref)
    for cell, listname in (("C5", "CityList"), ("G5", "CatList"),
                           ("K5", "CardList")):
        dv = DataValidation(type="list", formula1=f"={listname}", allow_blank=False)
        dash.add_data_validation(dv)
        dv.add(dash[cell])

    tile(dash, 2, "Total Spend (excl. payments)", f"=SUMPRODUCT({base}*{r_amt})", money=True)
    tile(dash, 5, "# Purchases", f"=SUMPRODUCT({base}*({r_amt}>0))", number=True)
    tile(dash, 8, "Avg Purchase",
         f"=IFERROR(SUMPRODUCT({base}*({r_amt}>0)*{r_amt})/SUMPRODUCT({base}*({r_amt}>0)),0)", money=True)
    tile(dash, 11, "Median Purchase",
         ArrayFormula("K10", f"=MEDIAN(IF({base}*({r_amt}>0),{r_amt}))"), money=True)

    dash.add_chart(pie_chart("Spend by Category", data, *T["cat"], width=16), "B12")
    dash.add_chart(column_chart("Spend by Month", data, *T["month"], width=16), "J12")
    dash.add_chart(column_chart("Spend by Day of Week", data, *T["dow"], width=16), "R12")
    dash.add_chart(column_chart("Spend by City", data, *T["city"], width=16), "B30")
    dash.add_chart(column_chart("Top Merchants", data, *T["merch"], width=16), "J30")
    dash.add_chart(pie_chart("Spend by Card", data, *T["card"], width=16), "R30")
    dash.add_chart(line_chart("Spend Over Time", data, *T["trend"], width=32), "B48")
    dash.add_chart(column_chart("Transaction Size Distribution", data, *T["size"], width=16), "R48")

    save(wb, len(df), dupes_removed, out or OUTPUT_FILE)


def save(wb, n, dupes_removed, target=OUTPUT_FILE):
    try:
        wb.save(target)
    except PermissionError:
        target = OUTPUT_FILE.with_name(f"expense_tracker_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
        wb.save(target)
        print(f"  ! {OUTPUT_FILE.name} was open in Excel - wrote {target.name} instead.")
    print(f"Done: {n} transactions, {dupes_removed} duplicates removed -> {target.name}")


def main():
    try:
        df = load_all()
    except (FileNotFoundError, ValueError) as exc:
        print(f"Nothing to do: {exc}")
        return 1
    df, dupes_removed = dedupe(df)
    rules = json.loads(CATEGORIES_FILE.read_text())
    mmap = load_merchant_map()
    df = enrich_dates(df)
    df = apply_categories(df, rules, mmap)
    mmap = prompt_new_merchants(df, rules, mmap)
    df = apply_categories(df, rules, mmap)
    build_workbook(df, dupes_removed)
    return 0


if __name__ == "__main__":
    sys.exit(main())
